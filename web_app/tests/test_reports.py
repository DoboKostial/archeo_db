from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.queries import insert_photogram_sql, report_photograms_table_list_all_sql, update_photogram_sql
from app.reports.context import ReportContext
from app.reports.exporters import geopts_table, photograms_table, photos_table
from app.reports.exporters.registry import EXPORTERS
from app.reports.exporters.utils_sql import sql_quote
from app.reports.registry import REPORT_GENERATORS, REPORT_SPECS, ReportSpec
from app.reports.service import init_report_generators
from app.routes import reports as report_routes


class _Cursor:
    def __init__(self, columns, rows):
        self.description = [(column,) for column in columns]
        self.rows = rows
        self.executed = []

    def __enter__(self):
        return self

    def __exit__(self, _exc_type, _exc, _tb):
        return False

    def execute(self, query, params=None):
        self.executed.append((query, params))

    def fetchall(self):
        return self.rows


class _Connection:
    def __init__(self, columns, rows):
        self.cursor_obj = _Cursor(columns, rows)

    def __enter__(self):
        return self

    def __exit__(self, _exc_type, _exc, _tb):
        return False

    def cursor(self):
        return self.cursor_obj


def _ctx():
    return ReportContext(
        lang="en",
        locale="en_US",
        selected_db="02_test",
        user_email="tester@example.invalid",
        t=lambda key: key,
    )


def _select_test_db(client):
    with client.session_transaction() as session:
        session["selected_db"] = "02_test"


def test_sql_quote_handles_json_dates_bytes_and_backslashes():
    assert sql_quote({"present": True, "note": "O'Reilly"}) == "'{\"present\": true, \"note\": \"O''Reilly\"}'"
    assert sql_quote(["alpha", 2]) == "'[\"alpha\", 2]'"
    assert sql_quote(date(2026, 8, 7)) == "'2026-08-07'"
    assert sql_quote(datetime(2026, 8, 7, 9, 30, tzinfo=timezone.utc)) == "'2026-08-07 09:30:00+00:00'"
    assert sql_quote(Decimal("12.50")) == "12.50"
    assert sql_quote(b"\x00\xff") == "decode('00ff', 'hex')"
    assert sql_quote(r"C:\field\photo") == r"'C:\field\photo'"


def test_geopts_sql_export_excludes_derived_geometry(monkeypatch):
    conn = _Connection(
        ["id_pts", "x", "y", "h", "code", "notes"],
        [(1, 10.0, 20.0, 30.0, "SU", "O'Reilly")],
    )
    monkeypatch.setattr(geopts_table, "get_terrain_connection", lambda _dbname: conn)

    sql_text = geopts_table.GeoptsTableExporter().to_sql(_ctx())

    assert "pts_geom" not in sql_text
    assert "INSERT INTO tab_geopts (id_pts, x, y, h, code, notes)" in sql_text
    assert "'O''Reilly'" in sql_text


def test_photos_sql_export_excludes_centroid_and_keeps_json(monkeypatch):
    monkeypatch.setattr(
        photos_table.PhotosTableExporter,
        "_fetch_rows",
        lambda _self, _ctx: [{"id_photo": "1_photo.jpg"}],
    )
    conn = _Connection(
        [
            "id_photo", "photo_typ", "datum", "author", "notes",
            "mime_type", "file_size", "checksum_sha256",
            "shoot_datetime", "gps_lat", "gps_lon", "gps_alt",
            "exif_json",
        ],
        [(
            "1_photo.jpg", "overview", date(2026, 8, 7), "author",
            "note", "image/jpeg", 123, "abc", None, 50.0, 14.0, 250.0,
            {"camera": "UnitCam", "ok": True},
        )],
    )
    monkeypatch.setattr(photos_table, "get_terrain_connection", lambda _dbname: conn)

    sql_text = photos_table.PhotosTableExporter().to_sql(_ctx())

    assert "photo_centroid" not in sql_text
    assert "exif_json" in sql_text
    assert '"ok": true' in sql_text


def test_photogram_date_is_part_of_write_and_report_queries():
    assert "photogram_typ, datum" in insert_photogram_sql()
    assert "datum=%s" in update_photogram_sql()
    assert "p.datum" in report_photograms_table_list_all_sql()


def test_photogram_xlsx_exports_date(monkeypatch):
    monkeypatch.setattr(
        photograms_table.PhotogramsTableExporter,
        "_fetch_rows",
        lambda _self, _ctx: [{
            "id_photogram": "1_model.jpg",
            "photogram_typ": "synthetic",
            "datum": date(2026, 8, 15),
        }],
    )
    monkeypatch.setattr(photograms_table, "list_files_for_media_id", lambda *_args: [])

    content = photograms_table.PhotogramsTableExporter().to_xlsx(_ctx())
    sheet = load_workbook(BytesIO(content)).active

    headers = [cell.value for cell in sheet[1]]
    assert "datum" in headers
    assert sheet.cell(row=2, column=headers.index("datum") + 1).value == datetime(2026, 8, 15)


def test_report_registry_matches_generators_and_exporters():
    init_report_generators()

    assert sorted(set(REPORT_SPECS) - set(REPORT_GENERATORS)) == []
    assert sorted(set(REPORT_SPECS) - set(EXPORTERS)) == []
    assert sorted(set(EXPORTERS) - set(REPORT_SPECS)) == []


def test_reports_fallback_keeps_selected_database(client, monkeypatch):
    _select_test_db(client)
    monkeypatch.setattr(
        report_routes.translator,
        "get_language_specs",
        lambda: (_ for _ in ()).throw(RuntimeError("boom")),
    )

    response = client.get("/reports")

    assert response.status_code == 200
    assert "02_test" in response.get_data(as_text=True)


def test_export_route_rejects_unsupported_format_before_exporter(client, monkeypatch):
    _select_test_db(client)
    monkeypatch.setitem(
        report_routes.REPORT_SPECS,
        "xlsx_only",
        ReportSpec(
            report_id="xlsx_only",
            title_key="report.xlsx_only.title",
            description_key="report.xlsx_only.description",
            formats=frozenset({"xlsx"}),
        ),
    )
    monkeypatch.setattr(
        report_routes,
        "get_exporter",
        lambda _export_id: (_ for _ in ()).throw(AssertionError("exporter should not be called")),
    )

    response = client.post("/reports/export/sql/xlsx_only")

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/reports?lang=en")
