# app/reports/exporters/polygon_cards.py
from __future__ import annotations
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple
from openpyxl import Workbook

from app.logger import logger
from app.database import get_terrain_connection
from app.reports.context import ReportContext

from app.queries import (
    report_polygon_cards_list_polygons_sql,
    report_polygon_cards_detail_sql,
    report_polygon_cards_bindings_top_sql,
    report_polygon_cards_bindings_bottom_sql,
    report_polygon_cards_sj_ids_sql,
    report_polygon_cards_media_ids_sql,
)

from .utils_excel import set_basic_column_widths
from .utils_media import list_files_for_media_id
from .utils_sql import dump_table_inserts, dump_table_inserts_columns


class PolygonCardsExporter:
    export_id = "polygon_cards"

    def _fetch_polygon_names(self, ctx: ReportContext) -> List[str]:
        with get_terrain_connection(ctx.selected_db) as conn:
            with conn.cursor() as cur:
                cur.execute(report_polygon_cards_list_polygons_sql())
                return [str(r[0]) for r in cur.fetchall()]

    def _fetch_polygon_detail(self, conn, polygon_name: str) -> Dict[str, Any]:
        with conn.cursor() as cur:
            cur.execute(report_polygon_cards_detail_sql(), (polygon_name,))
            row = cur.fetchone()
            if not row:
                return {}
            cols = [d[0] for d in cur.description]
            return dict(zip(cols, row))

    # -------------------------
    # Excel
    # -------------------------
    def to_xlsx(self, ctx: ReportContext) -> bytes:
        names = self._fetch_polygon_names(ctx)
        logger.info(f"[{ctx.selected_db}] Export XLSX polygon_cards: {len(names)} polygons lang={ctx.lang}")

        wb = Workbook()

        ws = wb.active
        ws.title = "Polygons"

        headers = [
            "polygon_name", "parent_name", "allocation_reason", "notes",
            "has_geom_top", "has_geom_bottom",
            "srid", "npoints_top", "npoints_bottom", "npoints_total",
            "area_top_m2", "area_bottom_m2",
        ]
        ws.append(headers)

        ws_top = wb.create_sheet("BindingsTop")
        ws_top.append(["ref_polygon", "pts_from", "pts_to"])

        ws_bottom = wb.create_sheet("BindingsBottom")
        ws_bottom.append(["ref_polygon", "pts_from", "pts_to"])

        ws_sj = wb.create_sheet("SJLinks")
        ws_sj.append(["ref_polygon", "ref_sj"])

        ws_media = wb.create_sheet("MediaLinks")
        ws_media.append(["ref_polygon", "kind", "ref_media", "files"])

        with get_terrain_connection(ctx.selected_db) as conn:
            for poly in names:
                d = self._fetch_polygon_detail(conn, poly)
                if not d:
                    continue

                ws.append([
                    d.get("polygon_name"),
                    d.get("parent_name"),
                    d.get("allocation_reason"),
                    d.get("notes"),
                    d.get("has_geom_top"),
                    d.get("has_geom_bottom"),
                    d.get("srid"),
                    d.get("npoints_top"),
                    d.get("npoints_bottom"),
                    d.get("npoints_total"),
                    d.get("area_top_m2"),
                    d.get("area_bottom_m2"),
                ])

                with conn.cursor() as cur:
                    cur.execute(report_polygon_cards_bindings_top_sql(), (poly,))
                    for a, b in cur.fetchall():
                        ws_top.append([poly, a, b])

                    cur.execute(report_polygon_cards_bindings_bottom_sql(), (poly,))
                    for a, b in cur.fetchall():
                        ws_bottom.append([poly, a, b])

                    cur.execute(report_polygon_cards_sj_ids_sql(), (poly,))
                    for (sj_id,) in cur.fetchall():
                        ws_sj.append([poly, sj_id])

                for kind in ("photos", "sketches", "photograms"):
                    with conn.cursor() as cur:
                        cur.execute(report_polygon_cards_media_ids_sql(kind), (poly,))
                        mids = [str(r[0]) for r in cur.fetchall()]
                    for mid in mids:
                        files = ", ".join(list_files_for_media_id(ctx, kind, mid))
                        ws_media.append([poly, kind, mid, files])

        set_basic_column_widths(ws, headers)
        set_basic_column_widths(ws_top, ["ref_polygon", "pts_from", "pts_to"])
        set_basic_column_widths(ws_bottom, ["ref_polygon", "pts_from", "pts_to"])
        set_basic_column_widths(ws_sj, ["ref_polygon", "ref_sj"])
        set_basic_column_widths(ws_media, ["ref_polygon", "kind", "ref_media", "files"])

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # -------------------------
    # SQL
    # -------------------------
    def to_sql(self, ctx: ReportContext) -> str:
        names = self._fetch_polygon_names(ctx)
        logger.info(f"[{ctx.selected_db}] Export SQL polygon_cards: {len(names)} polygons")

        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        header = [
            f"-- ArcheoDB export: polygon_cards",
            f"-- Database: {ctx.selected_db}",
            f"-- Generated: {ts}",
            f"-- NOTE: data-only dump (INSERTs). No binaries, NO polygon geometries (geom_top/geom_bottom excluded).",
            "",
            "BEGIN;",
            "",
        ]

        if not names:
            return "\n".join(header + ["COMMIT;", ""])

        with get_terrain_connection(ctx.selected_db) as conn:
            with conn.cursor() as cur:
                out: List[str] = header[:]

                # tab_polygons WITHOUT geom columns
                out.append(dump_table_inserts_columns(
                    cur,
                    "tab_polygons",
                    ["polygon_name", "parent_name", "allocation_reason", "notes"],
                    "WHERE polygon_name = ANY(%s)",
                    (names,),
                ))

                out.append(dump_table_inserts(
                    cur, "tab_polygon_geopts_binding_top",
                    "WHERE ref_polygon = ANY(%s)",
                    (names,),
                ))
                out.append(dump_table_inserts(
                    cur, "tab_polygon_geopts_binding_bottom",
                    "WHERE ref_polygon = ANY(%s)",
                    (names,),
                ))

                out.append(dump_table_inserts(
                    cur, "tabaid_sj_polygon",
                    "WHERE ref_polygon = ANY(%s)",
                    (names,),
                ))

                out.append(dump_table_inserts(
                    cur, "tabaid_polygon_photos",
                    "WHERE ref_polygon = ANY(%s)",
                    (names,),
                ))
                out.append(dump_table_inserts(
                    cur, "tabaid_polygon_photograms",
                    "WHERE ref_polygon = ANY(%s)",
                    (names,),
                ))
                out.append(dump_table_inserts(
                    cur, "tabaid_polygon_sketches",
                    "WHERE ref_polygon = ANY(%s)",
                    (names,),
                ))

                out.append("COMMIT;\n")
                return "\n".join(s for s in out if s)