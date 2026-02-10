# app/reports/exporters.py
from __future__ import annotations

import os
from io import BytesIO
from datetime import datetime
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.logger import logger
from app.database import get_terrain_connection
from app.reports.context import ReportContext
from config import Config

from app.queries import (
    report_sj_cards_list_sj_sql,
    report_sj_cards_detail_sql,
    report_sj_cards_media_ids_sql,
)


MEDIA_KINDS = ("photos", "drawings", "sketches", "photograms")


def _media_dir(ctx: ReportContext, kind: str) -> str:
    sub = (Config.MEDIA_DIRS or {}).get(kind, "")
    if not sub:
        return ""
    return os.path.join(Config.DATA_DIR, ctx.selected_db, sub)


def _list_files_for_media_id(ctx: ReportContext, kind: str, media_id: str) -> List[str]:
    """
    Return filenames (not thumbnails) for media_id in:
      DATA_DIR/<db>/<kind>/
    We list files that start with '<id>.' (any ext).
    """
    base = _media_dir(ctx, kind)
    if not base or not os.path.isdir(base):
        return []

    out: List[str] = []
    prefix = f"{media_id}."
    try:
        for fn in os.listdir(base):
            if fn.startswith(prefix):
                out.append(fn)
    except Exception:
        return []

    return sorted(out)


def _fetch_sj_ids(ctx: ReportContext) -> List[int]:
    with get_terrain_connection(ctx.selected_db) as conn:
        with conn.cursor() as cur:
            cur.execute(report_sj_cards_list_sj_sql())
            return [int(r[0]) for r in cur.fetchall()]


def _fetch_sj_detail(conn, sj_id: int) -> Dict[str, Any]:
    with conn.cursor() as cur:
        cur.execute(report_sj_cards_detail_sql(), (sj_id,))
        row = cur.fetchone()
        if not row:
            return {}
        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


def _fetch_media_ids(conn, sj_id: int, kind: str) -> List[str]:
    with conn.cursor() as cur:
        cur.execute(report_sj_cards_media_ids_sql(kind), (sj_id,))
        return [str(r[0]) for r in cur.fetchall()]


# -------------------------
# Excel export
# -------------------------

def export_sj_cards_excel(ctx: ReportContext) -> bytes:
    sj_ids = _fetch_sj_ids(ctx)
    logger.info(f"[{ctx.selected_db}] Export Excel SJ cards: {len(sj_ids)} SJs lang={ctx.lang}")

    wb = Workbook()
    ws = wb.active
    ws.title = "SJs"

    headers = [
        "id_sj", "sj_typ", "sj_subtype",
        "author", "recorded",
        "docu_plan", "docu_vertical", "ref_object",
        "description", "interpretation",
        "strat_above", "strat_below", "strat_equal",
        # subtype compact (deposit)
        "deposit_typ", "deposit_color", "deposit_boundary_visibility", "deposit_structure", "deposit_compactness", "deposit_removed",
        # subtype compact (negativ)
        "negativ_typ", "negativ_excav_extent", "negativ_ident_niveau_cut", "negativ_shape_plan", "negativ_shape_sides", "negativ_shape_bottom",
        # subtype compact (structure)
        "structure_typ", "structure_construction_typ", "structure_binder", "structure_basic_material", "structure_length_m", "structure_width_m", "structure_height_m",
        # media file lists (filenames only)
        "photos_files", "drawings_files", "sketches_files", "photograms_files",
    ]
    ws.append(headers)

    with get_terrain_connection(ctx.selected_db) as conn:
        for sj_id in sj_ids:
            sj = _fetch_sj_detail(conn, sj_id)
            if not sj:
                continue

            media_files: Dict[str, str] = {}
            for kind in MEDIA_KINDS:
                ids = _fetch_media_ids(conn, sj_id, kind)
                files: List[str] = []
                for mid in ids:
                    files.extend(_list_files_for_media_id(ctx, kind, mid))
                # store as comma-separated filenames
                media_files[kind] = ", ".join(files)

            row = [
                sj.get("id_sj"),
                sj.get("sj_typ"),
                sj.get("sj_subtype"),
                sj.get("author"),
                sj.get("recorded"),
                sj.get("docu_plan"),
                sj.get("docu_vertical"),
                sj.get("ref_object"),
                sj.get("description"),
                sj.get("interpretation"),
                sj.get("strat_above"),
                sj.get("strat_below"),
                sj.get("strat_equal"),

                sj.get("deposit_typ"),
                sj.get("deposit_color"),
                sj.get("deposit_boundary_visibility"),
                sj.get("deposit_structure"),
                sj.get("deposit_compactness"),
                sj.get("deposit_removed"),

                sj.get("negativ_typ"),
                sj.get("negativ_excav_extent"),
                sj.get("negativ_ident_niveau_cut"),
                sj.get("negativ_shape_plan"),
                sj.get("negativ_shape_sides"),
                sj.get("negativ_shape_bottom"),

                sj.get("structure_typ"),
                sj.get("structure_construction_typ"),
                sj.get("structure_binder"),
                sj.get("structure_basic_material"),
                sj.get("structure_length_m"),
                sj.get("structure_width_m"),
                sj.get("structure_height_m"),

                media_files["photos"],
                media_files["drawings"],
                media_files["sketches"],
                media_files["photograms"],
            ]
            ws.append(row)

    # Simple column width auto-fit (cheap)
    for col_idx in range(1, len(headers) + 1):
        col = get_column_letter(col_idx)
        ws.column_dimensions[col].width = min(60, max(12, len(headers[col_idx - 1]) + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -------------------------
# SQL export (INSERT dump)
# -------------------------

def _sql_quote(v: Any) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, (datetime, )):
        return "'" + v.isoformat(sep=" ", timespec="seconds") + "'"
    s = str(v)
    s = s.replace("\\", "\\\\").replace("'", "''")
    return f"'{s}'"


def _dump_table_inserts(cur, table: str, where_sql: str, params: Tuple[Any, ...]) -> str:
    """
    Dump rows from table as INSERT statements using SELECT * column order.
    """
    cur.execute(f"SELECT * FROM {table} {where_sql};", params)
    rows = cur.fetchall()
    if not rows:
        return ""

    cols = [d[0] for d in cur.description]
    cols_sql = ", ".join(cols)

    out_lines: List[str] = []
    for r in rows:
        values_sql = ", ".join(_sql_quote(v) for v in r)
        out_lines.append(f"INSERT INTO {table} ({cols_sql}) VALUES ({values_sql});")
    return "\n".join(out_lines) + "\n"


def export_sj_cards_sql(ctx: ReportContext) -> str:
    sj_ids = _fetch_sj_ids(ctx)
    logger.info(f"[{ctx.selected_db}] Export SQL SJ cards: {len(sj_ids)} SJs")

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    header = [
        f"-- ArcheoDB export: SJ cards",
        f"-- Database: {ctx.selected_db}",
        f"-- Generated: {ts}",
        f"-- NOTE: This is data-only dump (INSERTs). No binaries included.",
        "",
        "BEGIN;",
        "",
    ]

    if not sj_ids:
        return "\n".join(header + ["COMMIT;", ""])

    # We export base + 1:1 subtype + stratigraphy involving exported SJs + media link tables + finds/samples
    id_tuple = tuple(sj_ids)

    with get_terrain_connection(ctx.selected_db) as conn:
        with conn.cursor() as cur:
            out: List[str] = header[:]

            # tab_sj
            out.append(_dump_table_inserts(cur, "tab_sj", "WHERE id_sj = ANY(%s)", (list(id_tuple),)))

            # type-specific 1:1
            out.append(_dump_table_inserts(cur, "tab_sj_deposit", "WHERE id_deposit = ANY(%s)", (list(id_tuple),)))
            out.append(_dump_table_inserts(cur, "tab_sj_negativ", "WHERE id_negativ = ANY(%s)", (list(id_tuple),)))
            out.append(_dump_table_inserts(cur, "tab_sj_structure", "WHERE id_structure = ANY(%s)", (list(id_tuple),)))

            # stratigraphy: any relation where either side is in export set
            out.append(_dump_table_inserts(
                cur,
                "tab_sj_stratigraphy",
                "WHERE ref_sj1 = ANY(%s) OR ref_sj2 = ANY(%s)",
                (list(id_tuple), list(id_tuple)),
            ))

            # finds/samples
            out.append(_dump_table_inserts(cur, "tab_finds", "WHERE ref_sj = ANY(%s)", (list(id_tuple),)))
            out.append(_dump_table_inserts(cur, "tab_samples", "WHERE ref_sj = ANY(%s)", (list(id_tuple),)))

            # media link tables (data-only; media binary files are outside DB)
            out.append(_dump_table_inserts(cur, "tabaid_photo_sj", "WHERE ref_sj = ANY(%s)", (list(id_tuple),)))
            out.append(_dump_table_inserts(cur, "tabaid_sj_drawings", "WHERE ref_sj = ANY(%s)", (list(id_tuple),)))
            out.append(_dump_table_inserts(cur, "tabaid_sj_sketch", "WHERE ref_sj = ANY(%s)", (list(id_tuple),)))
            out.append(_dump_table_inserts(cur, "tabaid_photogram_sj", "WHERE ref_sj = ANY(%s)", (list(id_tuple),)))

            out.append("COMMIT;\n")
            return "\n".join(s for s in out if s)  # drop empty chunks
